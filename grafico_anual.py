import os
from datetime import datetime
import pyperclip
import locale
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill, Color
from openpyxl.worksheet import page

NOMBRE_TXT = 'turnos.txt'
NOMBRE_EXCEL = 'grafico_maquinistas_31-08_31-12.xlsx'
AGENTES = [
    '677, OCHOA BAEZA, ANTONIO',
    '2560, SANJUAN IZQUIERDO, YOLANDA',
    '2082, DE LA TORRE HERRERA, RICARDO',
    '109, BIELSA NOGUES, ERNESTO',
    '2049, COLLADO MASCAROS, IVAN',
    '409, GOMEZ SEGURA, JOSE',
    '2154, LARA NAVARRO, MANUEL',
    '1651, PASTOR BERTOMEU, JAIME JUAN',
    '2076, MENDOZA RODRIGUEZ, JOSE ANTONIO',
    '597, MENA DIAZ, VICTOR',
    '2316, ESTEVE GIMENO, ENRIQUETA',
    '2586, PEREZ SANTAMARIA, MONICA',
    '159, CARMONA GARCIA, DIEGO JOSE',
    '133, BOU SERRALTA, MIGUEL ANGEL',
    '2588, LLOPES MOLINA, JUAN MANUEL',
    '2369, CABRERA GUERRERO, ENRIQUE',
    '1204, DE LAMA GONZALEZ, CARLOS',
    '1472, GOMEZ MARTIN, JAIME',
    '2210, MORALES ASENSI, ENRIQUE']
NUM_AGENTES = len(AGENTES)
MESES = [
    'JULIO 2020',
    'AGOSTO 2020',
    'SEPTIEMBRE 2020',
    'OCTUBRE 2020',
    'NOVIEMBRE 2020',
    'DICIEMBRE 2020']
FESTIVOS = [
    '9-10-2020',
    '12-10-2020',
    '8-12-2020',
    '25-12-2020']
TURNOS_M = [1, 2, 3, 4, 8]
TURNOS_T = [5, 6, 7, 9]
TURNOS_D = ['D', 'DT']
TURNOS_FG = ['FGV', 'FGI', 'FO']

## ESTILOS ##
negrita = Font(bold=True)
centrado = Alignment(horizontal='center', vertical='center')
titulo = NamedStyle(
    name='titulo',
    font=Font(bold=True, size=18),
    alignment=Alignment(horizontal='center', vertical='center'))
estilo_mes = NamedStyle(
    name='dia_mes',
    number_format='dd mmm',
    font=Font(bold=True, size=11),
    alignment=Alignment(horizontal='center', vertical='center'))
estilo_mes_rojo = NamedStyle(
    name='dia_mes_rojo',
    number_format='dd mmm',
    font=Font(bold=True, size=11, color='FF2600'),
    alignment=Alignment(horizontal='center', vertical='center'))
estilo_sem = NamedStyle(
    name='dia_sem',
    number_format='ddd',
    font=Font(bold=True, size=11),
    alignment=Alignment(horizontal='center', vertical='center'))
estilo_sem_rojo = NamedStyle(
    name='dia_sem_rojo',
    number_format='ddd',
    font=Font(bold=True, size=11, color='FF2600'),
    alignment=Alignment(horizontal='center', vertical='center'))
fondo_gris = NamedStyle(
    name='fondo_gris',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color('CBCBCB')))

# Colores turnos
ROJO = 'E38BAF'
VERDE = 'C8F0BD'
AMARILLO = 'FFFFA6'
AZUL = '90D1DB'
NARANJA = 'FFB940'
manyana = NamedStyle(
    name='manyana',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color(AMARILLO)))
tarde = NamedStyle(
    name='tarde',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color(AZUL)))
festivo = NamedStyle(
    name='festivo',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color(VERDE)))
fg = NamedStyle(
    name='fg',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color(NARANJA)))
vacaciones = NamedStyle(
    name='vacaciones',
    fill=PatternFill(
        patternType='solid',
        fill_type='solid',
        fgColor=Color(ROJO)))

# Para que pueda detectar el nombre de los meses en español
locale.setlocale(locale.LC_ALL, 'es_ES.utf-8')


def pdf_a_txt():
    """
    Crea un archivo de texto con las hojas del pdf que se van copiando
    al portapapeles una a una. Al finalizar, en las pruebas realizadas,
    el fichero resultante no ha sido perfecto, teniendo que ejecutar la
    funcion lee_turnos varias veces para localizar los fallos y corregirlos
    manualmente. Los fallos han sido saltos de línea que no correspondían
    con el final del mes, y espacios en blanco. En ambos casos ha bastado
    con hacer una busqueda y sustitución por comas.
    """
    with open(NOMBRE_TXT, 'a') as file_obj:
        for mes in MESES:
            input(f'Copia el mes {mes} y pulsa Enter')
            texto = pyperclip.paste()
            texto = texto.replace('\n', ',')
            file_obj.write(texto + '\n')


def txt_a_excel():
    """
    Lee el contenido del fichero turnos.txt y crea una lista con un diccionario
    por cada día del mes, que incluyen la fecha y una lista con los turnos de
    esa fecha.

    Cuando se ha creado la lista con todos los turnos del mes, se llama a la
    función crea_excel, pasando esta lista como parámetro.
    """
    meses = (mes for mes in open(NOMBRE_TXT))
    for nombre_mes in MESES:
        lista_dias = []
        mes_gen = (s.replace('\n', '').split(',') for s in meses)
        mes = next(mes_gen)
        for i in range(0, len(mes), NUM_AGENTES + 2):
            d = mes[i:i + NUM_AGENTES + 2]
            fecha = datetime.strptime(d[0].replace('.', '-2020'), '%d-%b-%Y')
            dia = {
                'fecha': fecha,
                'turnos': d[2:]
            }
            lista_dias.append(dia)
        crea_excel(AGENTES, lista_dias, nombre_mes)


def crea_excel(agentes, dias, mes):
    """
    Crea un fichero excel con los datos obtenidos anteriormente.

        Parametros:
            agentes (list) : Lista de los agentes.
            dias (list) : Lista con diccionarios de cada día del mes.
            mes (str) : Mes y año.
    """

    # Si ya existe el fichero lo abre.
    if os.path.isfile(NOMBRE_EXCEL):
        wb = openpyxl.load_workbook(NOMBRE_EXCEL)
    # En caso de que no exista lo crea.
    else:
        wb = openpyxl.Workbook()

    # Crea una nueva hoja con el nombre del mes y año.
    wb.create_sheet(title=mes, index=MESES.index(mes))

    # Borra la hoja inicial 'Sheet'
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Hoja del mes actual
    sheet = wb[mes]

    # Tamaño y orientación de la página
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.paperWidth = '170mm'
    sheet.page_setup.paperHeight = '520mm'
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    
    # Título de la hoja (mes y año)
    sheet['A1'] = mes

    # Introduce apellidos y nombre de los agentes en las columnas A y B
    for fila in range(3, 3 + len(AGENTES)):
        sheet.cell(row=fila, column=1).value = AGENTES[fila - 3].split(',')[1].lstrip()
        sheet.cell(row=fila, column=2).value = AGENTES[fila - 3].split(',')[2].lstrip()

    # Introduce los turnos del mes en columnas por día, utilizando la
    # primera y segunda fila para el día del mes, y el resto de filas para
    # los turnos de cada agente.
    for columna in range(3, 3 + len(dias)):
        dia = dias[columna - 3]
        fecha = dia['fecha']
        sheet.cell(row=1, column=columna).value = fecha
        sheet.cell(row=2, column=columna).value = fecha
        for fila in range(3, 3 + len(AGENTES)):
            if dia['turnos'][fila - 3].isdigit():
                sheet.cell(row=fila, column=columna).value = int(dia['turnos'][fila - 3])
            else:
                sheet.cell(row=fila, column=columna).value = dia['turnos'][fila - 3]

    # Graba los datos en el fichero excel.
    wb.save(NOMBRE_EXCEL)


def formatea_excel():
    """
    Da formato al fichero excel
    """
    # Comprueba si existe el fichero
    if not os.path.isfile(NOMBRE_EXCEL):
        print(f"No se encuentra el fichero excel {NOMBRE_EXCEL}.")
        sys.exit(1)
    else:
        wb = openpyxl.load_workbook(NOMBRE_EXCEL)
    
    for sheet in wb.sheetnames:
        # Tamaño de la columna para los apellidos
        wb[sheet].column_dimensions['A'].width = 23
        # Tamaño de la columna para el nombre
        wb[sheet].column_dimensions['B'].width = 15
        # Fusiona las celdas para el nombre del mes
        wb[sheet].merge_cells('A1:B2')
        wb[sheet]['A1'].style = titulo

        # Lineas impares en gris
        for fila in range(3, 3 + NUM_AGENTES):
            if fila % 2 != 0:
                for celda in wb[sheet][fila]:
                    celda.style = fondo_gris

        # Nombres y apellidos en negrita
        for fila in range(3, 3+ NUM_AGENTES):
            wb[sheet][f'A{fila}'].font = negrita
            wb[sheet][f'B{fila}'].font = negrita

        for columna in range(3, wb[sheet].max_column + 1):
            letra_col = get_column_letter(columna)
            # Tamaño de las columnas de los días
            wb[sheet].column_dimensions[letra_col].width = 7
            for fila in range(1, wb[sheet].max_row + 1):
                if fila == 1: # Día del mes
                    dia_mes = wb[sheet][f'{letra_col}{fila}']
                    if dia_mes.value.weekday() == 5 or dia_mes.value.weekday() == 6:
                        dia_mes.style = estilo_mes_rojo
                    else:
                        dia_mes.style = estilo_mes
                elif fila == 2: # Día de la semana
                    dia_semana = wb[sheet][f'{letra_col}{fila}']
                    if dia_semana.value.weekday() == 5 or dia_semana.value.weekday() == 6:
                        dia_semana.style = estilo_sem_rojo
                    else:
                        dia_semana.style = estilo_sem
                else:
                    wb[sheet][f'{letra_col}{fila}'].alignment = centrado

    wb.save(NOMBRE_EXCEL)


def pinta_festivos():
    """
    Recorre la lista FESTIVOS para pintar de rojo cada fecha.
    """
    # Comprueba si existe el fichero
    if not os.path.isfile(NOMBRE_EXCEL):
        print(f"No se encuentra el fichero excel {NOMBRE_EXCEL}.")
        sys.exit(1)
    else:
        wb = openpyxl.load_workbook(NOMBRE_EXCEL)

    for fecha in FESTIVOS:
        fecha = datetime.strptime(fecha, '%d-%m-%Y')
        nombre_hoja = fecha.strftime('%B %Y').upper()
        ws = wb[nombre_hoja]
        columna = int(fecha.strftime('%d')) + 2
        ws.cell(column=columna, row= 1).style = 'dia_mes_rojo'
        ws.cell(column=columna, row= 2).style = 'dia_sem_rojo'
        
    wb.save(NOMBRE_EXCEL)


def pinta_turnos():
    """
    Recorre todas las hojas y pinta los turnos con un color dependiendo de si es de
    mañana, tarde, festivo o FG.
    """
    # Comprueba si existe el fichero
    if not os.path.isfile(NOMBRE_EXCEL):
        print(f"\nERROR: No se encuentra el fichero excel {NOMBRE_EXCEL}.")
        exit(1)
    else:
        wb = openpyxl.load_workbook(NOMBRE_EXCEL)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rng = ws[f'C3:{get_column_letter(ws.max_column)}21']
        for fila in rng:
            for celda in fila:
                if celda.value in TURNOS_M:
                    celda.style = 'manyana'
                elif celda.value in TURNOS_T:
                    celda.style = 'tarde'
                elif celda.value in TURNOS_D:
                    celda.style = 'festivo'
                elif celda.value in TURNOS_FG:
                    celda.style = 'fg'
                celda.alignment = centrado
    wb.save(NOMBRE_EXCEL)


def main():
    # txt_a_excel()
    # formatea_excel()
    # pinta_festivos()
    pinta_turnos()


if __name__ == "__main__":
    main()
