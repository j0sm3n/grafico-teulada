import os
from datetime import datetime
import pyperclip
import locale
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, NamedStyle, Color, PatternFill


NOMBRE_EXCEL = 'test.xlsx'
AGENTES = [
    '677, OCHOA BAEZA, ANTONIO',
    '2560, SANJUAN IZQUIERDO, YOLANDA',
    '2082, DE LA TORRE HERRERA, RICARDO',
    '109, BIELSA NOGUES, ERNESTO',
    '2049, COLLADO MASCAROS, IVAN',
    '409, GOMEZ SEGURA, JOSE',
    '2154, LARA NAVARRO, MANUEL',
    '1651, PASTOR BERTOMEU, JAIME JUAN',
    '2076, MENDOZA RODRIGUEZ, JOSE ANTONIO',
    '597, MENA DIAZ, VICTOR',
    '2316, ESTEVE GIMENO, ENRIQUETA',
    '2586, PEREZ SANTAMARIA, MONICA',
    '159, CARMONA GARCIA, DIEGO JOSE',
    '133, BOU SERRALTA, MIGUEL ANGEL',
    '2588, LLOPES MOLINA, JUAN MANUEL',
    '2369, CABRERA GUERRERO, ENRIQUE',
    '1204, DE LAMA GONZALEZ, CARLOS',
    '1472, GOMEZ MARTIN, JAIME',
    '2210, MORALES ASENSI, ENRIQUE']
NUM_AGENTES = len(AGENTES)
MESES = [
    'JULIO 2020',
    'AGOSTO 2020',
    'SEPTIEMBRE 2020',
    'OCTUBRE 2020',
    'NOVIEMBRE 2020',
    'DICIEMBRE 2020']

## ESTILOS ##
negrita = Font(bold=True)
centrado = Alignment(horizontal='center', vertical='center')
titulo = NamedStyle(
    name='titulo',
    font=Font(bold=True, size=18),
    alignment=Alignment(horizontal='center', vertical='center'))
estilo_mes = NamedStyle(
        # name='dia_mes',
        number_format='dd mmm',
        font=Font(bold=True, size=11),
        alignment=Alignment(horizontal='center', vertical='center'))
estilo_mes_rojo = NamedStyle(
        # name='dia_mes_rojo',
        number_format='dd mmm',
        font=Font(bold=True, size=11, color='FF2600'),
        alignment=Alignment(horizontal='center', vertical='center'))
estilo_sem = NamedStyle(
        name='dia_sem',
        number_format='ddd',
        font=Font(bold=True, size=11),
        alignment=Alignment(horizontal='center', vertical='center'))
estilo_sem_rojo = NamedStyle(
        name='dia_sem_rojo',
        number_format='ddd',
        font=Font(bold=True, size=11, color='FF2600'),
        alignment=Alignment(horizontal='center', vertical='center'))
fondo_gris = NamedStyle(
        name='gris',
        fill=PatternFill(
            patternType='solid',
            fill_type='solid',
            fgColor=Color('CBCBCB')))


def lee_portapapeles():
    """
    Lee el contenido del portapapeles y formatea el texto, creando una 
    lista 'dias' con el siguiente contenido:
    dias[0] -> {'fecha': '10-jul-2020'}
    dias[1] -> No se utiliza
    dias[2:] -> {'turnos': ['3', 'D', 'D', '7', '8'...]}
    """

    # Para que pueda detectar el nombre de los meses en español
    locale.setlocale(locale.LC_ALL, 'es_ES')
    
    # Capturamos el portapapeles y creamos una lista, quitando los saltos
    # de línea y los espacios en blanco.
    texto = pyperclip.paste()
    texto.replace('\n', '')
    texto_lista = texto.split()
    
    # Creamos una lista, que contiene diccionarios con la fecha y los turnos
    # de cada día.
    dias = []
    for i in range(0, len(texto_lista), NUM_AGENTES + 2):
        d = texto_lista[i:i + NUM_AGENTES + 2]
        fecha = d[0].replace('.', '-2020')
        # print(fecha + " -> ", end=' ')
        fecha = datetime.strptime(fecha, '%d-%b-%Y')
        # print(fecha.strftime('%d/%m/%Y'))
        dia = {
            'fecha': fecha,
            'turnos': d[2:]
        }
        # print(f"Fecha: {dia['fecha']}", end=" ")
        # print(f"Trunos: {dia['turnos']}")
        dias.append(dia)

    return dias


def crea_excel(agentes, dias, mes):
    """
    Crea un fichero excel con los datos obtenidos anteriormente.
    """

    # Si ya existe el fichero lo abre.
    if os.path.isfile(NOMBRE_EXCEL):
        wb = openpyxl.load_workbook(NOMBRE_EXCEL)
    # En caso de que no exista lo crea.
    else:
        wb = openpyxl.Workbook()

    # Crea una nueva hoja con el nombre del mes y año.
    wb.create_sheet(title=mes, index=MESES.index(mes))

    # Borra la hoja inicial 'Sheet'
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Hoja del mes actual
    sheet = wb[mes]

    # Título de la hoja (mes y año)
    sheet['A1'] = mes
    sheet['A1'].style = titulo
    # Ancho de la columna para los apellidos
    sheet.column_dimensions['A'].width = 18
    # Ancho de la columna para el nombre
    sheet.column_dimensions['B'].width = 12
    
    # Introduce apellidos y nombre de los agentes en las columnas A y B
    for fila in range(3, 3 + NUM_AGENTES):
        # Apellidos
        apellidos = sheet.cell(row=fila, column=1)
        apellidos.value = AGENTES[fila - 3].split(',')[1].lstrip()
        if fila % 2 != 0:
            apellidos.style = fondo_gris
        apellidos.font = negrita
        # Nombre
        nombre = sheet.cell(row=fila, column=2)
        nombre.value = AGENTES[fila - 3].split(',')[2].lstrip()
        if fila % 2 != 0:
            nombre.style = fondo_gris
        nombre.font = negrita

    # Introduce los turnos del mes en columnas por día, utilizando la
    # primera y segunda fila para el día del mes, y el resto de filas para
    # los turnos de cada agente.
    for columna in range(3, 3 + len(dias)):
        # Ancho de la columna para los turnos
        letra_col = get_column_letter(columna)
        sheet.column_dimensions[letra_col].width = 6

        # Día de la columna actual
        dia = dias[columna - 3]

        # Día del mes
        dia_mes = sheet.cell(row=1, column=columna)
        dia_mes.value = dia['fecha']
        # Si es sábado o domingo utiliza letra roja
        if dia_mes.value.weekday() == 5 or dia_mes.value.weekday() == 6:
            dia_mes.style = estilo_mes_rojo
        else:
            dia_mes.style = estilo_mes

        # Día de la semana
        dia_semana = sheet.cell(row=2, column=columna)
        dia_semana.value = dia['fecha']
        # Si es sábado o domingo utiliza letra roja
        if dia_semana.value.weekday() == 5 or dia_semana.value.weekday() == 6:
            dia_semana.style = estilo_sem_rojo
        else:
            dia_semana.style = estilo_sem

        # Turnos del día actual
        for fila in range(3, 3 + len(dia['turnos'])):
            turno = sheet.cell(row=fila, column=columna)
            if dia['turnos'][fila - 3].isdigit():
                turno.value = int(dia['turnos'][fila - 3])
            else:
                turno.value = dia['turnos'][fila - 3]
            if fila % 2 != 0:
                turno.style = fondo_gris
            turno.alignment = centrado

    # Fusiona las celdas para el nombre del mes
    sheet.merge_cells('A1:B2')

    # Graba los datos en el fichero excel.
    wb.save(NOMBRE_EXCEL)


def turnos_bd(agentes, dias):
    """
    Crea una base de datos con los turnos de cada agente
    """
    # conexion = sqlite3.connect('grafico_anual.db')
    # cursor = conexion.cursor()

    for dia in dias:
        for i in range(NUM_AGENTES):
            sql = "INSERT INTO turnos_agente VALUES("
            sql += "null, "
            sql += f"'{dia['turnos'][i]}', "
            sql += f"'{agentes[i].split(',')[0]}', "
            sql += f"'{dia['fecha']}'"
            print(sql)


    # conexion.commit()
    # conexion.close()

def main():
    # Si ya existe el fichero nos avisa y termina la ejecución.
    if os.path.isfile(NOMBRE_EXCEL):
        print(f'ERROR: El fichero {NOMBRE_EXCEL} ya existe.')
        exit(1)
    
    for mes in MESES:
        print(f"Copia el mes {mes} en el portapapeles y pulsa enter para continuar. Pulsa C para cancelar.")
        resp = input("> ")
        if resp == 'c' or resp == 'C':
            break
        else:
            dias = lee_portapapeles()
            crea_excel(AGENTES, dias, mes)
            # turnos_bd(AGENTES,dias)


if __name__ == "__main__":
    main()
